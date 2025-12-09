import os
import re
from datetime import datetime, timezone
from functools import wraps
from io import BytesIO

import pandas as pd
from fpdf import FPDF
from flask import (
    Flask,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pymongo import MongoClient, errors
from bson import ObjectId
from bson.errors import InvalidId
from zoneinfo import ZoneInfo

from data import (
    current_year,
    dignitaries,
    events,
    gallery_slides,
    hero_story,
    highlights,
    schedule,
    storyline,
)

IST = ZoneInfo("Asia/Kolkata")

ADDITIONAL_GALLERY_FILES = [
    ("images/home_img/maxresdefault.jpg", "Lakshadeepotsava Celebration"),
    ("images/home_img/Chariot-udupi-krishna-matha.jpg", "Temple Chariot at Udupi Krishna Matha"),
    ("images/home_img/FpkEW5iXEAEwWC2.jpg", "Procession of Devotees"),
    ("images/home_img/dfdce484948261df4827ab9218a87260.jpg", "Alankara of Sri Krishna"),
    (
        "images/home_img/WhatsApp Image 2025-11-19 at 11.53.41_f9b61d6a.jpg",
        "Laksha Deepotsava Sevaks",
    ),
]

ASSET_VERSION = os.environ.get("ASSET_VERSION", "20241120")
CACHE_MAX_AGE = int(os.environ.get("CACHE_MAX_AGE", "600"))

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "change-this-in-production")
app.config["ADMIN_USERNAME"] = os.environ.get("ADMIN_USERNAME", "admin")
app.config["ADMIN_PASSWORD"] = os.environ.get("ADMIN_PASSWORD", "brahatgeetha2025")
app.config["MONGO_URI"] = os.environ.get("MONGO_URI", "mongodb://localhost:27017")
app.config["MONGO_DB_NAME"] = os.environ.get("MONGO_DB_NAME", "krishna_event")

DEFAULT_COLLEGES = [
    "Dr. B. B. Hegde First Grade College, Kundapura",
    "Govinda Dasa College, Surathkal",
    "Sri Bhuvanendra College, Karkala",
    "MGM College, Udupi",
    "Bhandarkars' Arts & Science College, Kundapura",
]

DEFAULT_COURSES = [
    "BCA",
    "B.Com",
    "B.Sc",
    "B.A",
    "MBA",
    "MCA",
]

mongo_client = None
mongo_db = None

PHONE_PATTERN = re.compile(r"^[6-9]\d{9}$")

def get_db():
    """Return MongoDB database handle or None if unavailable."""
    global mongo_client, mongo_db
    if mongo_db is not None:
        return mongo_db

    try:
        mongo_client = MongoClient(
            app.config["MONGO_URI"],
            serverSelectionTimeoutMS=3000,
        )
        mongo_client.admin.command("ping")
        mongo_db = mongo_client[app.config["MONGO_DB_NAME"]]
    except errors.PyMongoError as exc:
        app.logger.error("MongoDB connection failed: %s", exc)
        mongo_db = None

    return mongo_db


def get_form_options():
    """Fetch college and course options from MongoDB."""
    db = get_db()
    if db is None:
        return DEFAULT_COLLEGES, DEFAULT_COURSES

    doc = db.meta.find_one({"_id": "form_options"}) or {}
    colleges = doc.get("colleges") or DEFAULT_COLLEGES
    courses = doc.get("courses") or DEFAULT_COURSES
    return colleges, courses


def save_form_options(colleges, courses):
    db = get_db()
    if db is None:
        return False

    db.meta.update_one(
        {"_id": "form_options"},
        {"$set": {"colleges": colleges, "courses": courses}},
        upsert=True,
    )
    return True


def fetch_registrations(search_query=None, college_filter=None):
    """Fetch all registrations, optionally filtered by search query and/or college."""
    db = get_db()
    if db is None:
        return []

    query = {}
    
    # College filter (exact match)
    if college_filter and college_filter.strip():
        query["college"] = college_filter.strip()
    
    # Search query (across multiple fields)
    if search_query and search_query.strip():
        search_regex = {"$regex": search_query.strip().lower(), "$options": "i"}
        search_conditions = {
            "$or": [
                {"name": search_regex},
                {"college": search_regex},
                {"course": search_regex},
                {"role": search_regex},
                {"phone": search_regex},
                {"email": search_regex},
            ]
        }
        # Combine college filter with search
        if college_filter and college_filter.strip():
            query = {"$and": [{"college": college_filter.strip()}, search_conditions]}
        else:
            query = search_conditions

    registrations = []
    for entry in db.registrations.find(query).sort("created_at", 1):
        entry["_id"] = str(entry["_id"])
        entry["formatted_created_at"] = format_timestamp(entry.get("created_at"))
        registrations.append(entry)
    return registrations


def admin_required(view_func):
    @wraps(view_func)
    def wrapper(*args, **kwargs):
        if not session.get("admin_authenticated"):
            flash("Please sign in to access the admin console.", "warning")
            return redirect(url_for("admin_login"))
        return view_func(*args, **kwargs)

    return wrapper


def db_unavailable_message():
    flash(
        "The registration database is currently unreachable. "
        "Please try again once connectivity is restored.",
        "danger",
    )


def format_timestamp(value, fmt="%d %b %Y · %I:%M %p"):
    if isinstance(value, datetime):
        if value.tzinfo is None:
            value = value.replace(tzinfo=timezone.utc)
        return value.astimezone(IST).strftime(fmt)
    return value or ""


@app.template_filter("format_phone")
def format_phone(phone: str) -> str:
    """Return the raw 10-digit phone number for display/exports.

    Handles legacy values stored with +91 prefixes by stripping them.
    """
    if not phone:
        return ""
    digits = re.sub(r"\D", "", str(phone))
    if len(digits) == 12 and digits.startswith("91"):
        digits = digits[2:]
    return digits


def normalize_phone(raw_phone):
    digits = re.sub(r"\D", "", raw_phone or "")

    if len(digits) == 12 and digits.startswith("91"):
        digits = digits[2:]
    elif len(digits) == 11 and digits.startswith("0"):
        digits = digits[1:]

    if PHONE_PATTERN.fullmatch(digits):
        # Store just the 10-digit number; display formatting handled elsewhere.
        return digits
    return None


def get_gallery_images():
    """Get all images from static/images folder."""
    image_dir = os.path.join(app.static_folder, "images")
    gallery_images = []

    if os.path.exists(image_dir):
        image_extensions = {".jpg", ".jpeg", ".png", ".webp", ".gif", ".svg"}

        for filename in os.listdir(image_dir):
            lower_name = filename.lower()
            if (
                "logo" in lower_name
                or "hero" in lower_name
                or "pm modi-30oct25" in lower_name
                or "png-clipart-happy-krishna-janmashtami" in lower_name
                or "krishna-alankara-3.webp" in lower_name
                or "main_image.jpg" in lower_name
                or "udupi-krishna-idol.jpg" in lower_name
                or "udupi-krishna.webp" in lower_name
                or "yogi-sants-gathering.jpg" in lower_name
            ):
                continue

            _, ext = os.path.splitext(filename.lower())
            if ext in image_extensions:
                caption = (
                    filename.replace("-", " ")
                    .replace("_", " ")
                    .replace(ext, "")
                )
                caption = " ".join(word.capitalize() for word in caption.split())

                gallery_images.append(
                    {
                        "src": url_for("static", filename=f"images/{filename}"),
                        "caption": caption,
                        "title": caption,
                        "subtitle": "Sri Krishna Math · Devotional Moment",
                    }
                )

    return gallery_images if gallery_images else gallery_slides


def get_additional_gallery_images():
    images = []
    for filename, caption in ADDITIONAL_GALLERY_FILES:
        images.append(
            {
                "src": url_for("static", filename=filename),
                "caption": caption,
                "title": caption,
                "loading": "eager",
                "fetchpriority": "high",
            }
        )
    return images


@app.context_processor
def inject_layout_tokens():
    return {
        "brand": "Laksha Kantha Geetha Parayana",
        "year": current_year(),
        "nav_links": [
            {"href": "/", "label": "Home"},
            {"href": "/gallery", "label": "Gallery"},
            {"href": "/about", "label": "About"},
            {"href": "/register", "label": "Register"},
        ],
        "asset_version": ASSET_VERSION,
    }


@app.after_request
def apply_response_headers(response):
    response.headers["Cache-Control"] = f"public, max-age={CACHE_MAX_AGE}"
    response.headers.pop("Expires", None)
    response.headers.pop("X-Frame-Options", None)
    response.headers.pop("X-XSS-Protection", None)
    csp = (
        "default-src 'self'; "
        "img-src 'self' data: https://images.unsplash.com; "
        "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com https://cdnjs.cloudflare.com; "
        "font-src 'self' https://fonts.gstatic.com https://cdnjs.cloudflare.com data:; "
        "script-src 'self' 'unsafe-inline'; "
        "connect-src 'self'; "
        "frame-ancestors 'self';"
    )
    response.headers["Content-Security-Policy"] = csp
    return response

@app.route("/favicon.ico")
def favicon():
    # Return 204 No Content to remove favicon/logo from title
    from flask import Response
    return Response(status=204)

@app.route("/")
def home():
    return render_template(
        "home.html",
        hero=hero_story,
        highlights=highlights,
        schedule=schedule,
        events=events,
        dignitaries=dignitaries,
        storyline=storyline,
        gallery_preview=gallery_slides[:3],
    )


@app.route("/gallery")
def gallery():
    gallery_images = get_gallery_images() + get_additional_gallery_images()
    return render_template("gallery.html", hero=hero_story, gallery=gallery_images)


@app.route("/about")
def about():
    return render_template(
        "about.html",
        hero=hero_story,
        storyline=storyline,
        dignitaries=dignitaries,
    )


@app.route("/api/events")
def api_events():
    return jsonify(events)


@app.route("/register", methods=["GET", "POST"])
def register():
    colleges, courses = get_form_options()
    db = get_db()

    if request.method == "POST":
        raw_phone = request.form.get("phone", "").strip()
        form_data = {
            "name": request.form.get("name", "").strip(),
            "college": request.form.get("college", "").strip(),
            "course": request.form.get("course", "").strip(),
            "role": request.form.get("role", "").strip(),
            "email": request.form.get("email", "").strip(),
            "created_at": datetime.now(timezone.utc),
        }

        errors_list = []
        if not form_data["name"]:
            errors_list.append("Please enter your full name.")
        if not form_data["college"]:
            errors_list.append("Please select your college.")
        if not form_data["course"]:
            errors_list.append("Please select your course.")
        valid_roles = {"Student", "Faculty", "Volunteer"}
        if form_data["role"] not in valid_roles:
            errors_list.append("Please choose a valid role.")
        normalized_phone = normalize_phone(raw_phone)
        if not normalized_phone:
            errors_list.append("Please provide a valid 10-digit Indian mobile number.")
        if form_data["email"] and "@" not in form_data["email"]:
            errors_list.append("Please provide a valid email address.")

        if errors_list:
            for issue in errors_list:
                flash(issue, "danger")
            return render_template(
                "register.html",
                hero=hero_story,
                colleges=colleges,
                courses=courses,
            )

        if db is None:
            db_unavailable_message()
            return (
                render_template(
                    "register.html",
                    hero=hero_story,
                    colleges=colleges,
                    courses=courses,
                ),
                503,
            )

        form_data["phone"] = normalized_phone

        duplicate_phone = db.registrations.find_one(
            {"phone": {"$in": [normalized_phone, f"+91{normalized_phone}"]}}
        )
        duplicate_email = (
            db.registrations.find_one({"email": form_data["email"]})
            if form_data["email"]
            else None
        )

        duplicate_messages = []
        if duplicate_phone:
            duplicate_messages.append("This mobile number is already registered.")
        if duplicate_email:
            duplicate_messages.append("This email address is already registered.")

        if duplicate_messages:
            for msg in duplicate_messages:
                flash(msg, "warning")
            return render_template(
                "register.html",
                hero=hero_story,
                colleges=colleges,
                courses=courses,
                registration_duplicate=True,
            )

        db.registrations.insert_one(form_data)
        session["registration_success"] = True
        flash("Jai Sri Krishna! Your registration is confirmed.", "success")
        return redirect(url_for("register"))

    registration_success = session.pop("registration_success", False)
    registration_duplicate = session.pop("registration_duplicate", False)

    return render_template(
        "register.html",
        hero=hero_story,
        colleges=colleges,
        courses=courses,
        registration_success=registration_success,
        registration_duplicate=registration_duplicate,
    )


@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        if (
            username == app.config["ADMIN_USERNAME"]
            and password == app.config["ADMIN_PASSWORD"]
        ):
            session["admin_authenticated"] = True
            flash("Welcome back, administrator.", "success")
            return redirect(url_for("admin_dashboard"))
        flash("Invalid credentials. Try again.", "danger")

    return render_template("admin_login.html", hero=hero_story)


@app.route("/admin/logout")
def admin_logout():
    session.pop("admin_authenticated", None)
    flash("You have been signed out.", "info")
    return redirect(url_for("admin_login"))


@app.route("/admin/api/registrations", methods=["GET"])
def admin_api_registrations():
    """API endpoint for paginated and searchable registrations."""
    # Check authentication for API endpoint
    if not session.get("admin_authenticated"):
        return jsonify({"error": "Authentication required"}), 401
    
    db = get_db()
    if db is None:
        return jsonify({"error": "Database unavailable"}), 503

    # Get query parameters
    page = int(request.args.get("page", 1))
    limit = int(request.args.get("limit", 10))
    search_query = request.args.get("search", "").strip().lower()
    college_filter = request.args.get("college", "").strip()

    # Build query
    query = {}
    
    # College filter (exact match)
    if college_filter:
        query["college"] = college_filter
    
    # Search query (across multiple fields)
    if search_query:
        search_regex = {"$regex": search_query, "$options": "i"}
        search_conditions = {
            "$or": [
                {"name": search_regex},
                {"college": search_regex},
                {"course": search_regex},
                {"role": search_regex},
                {"phone": search_regex},
                {"email": search_regex},
            ]
        }
        # Combine college filter with search
        if college_filter:
            query = {"$and": [{"college": college_filter}, search_conditions]}
        else:
            query = search_conditions

    # Get total count
    total_count = db.registrations.count_documents(query)

    # Calculate skip
    skip = (page - 1) * limit

    # Fetch registrations
    registrations = []
    for entry in (
        db.registrations.find(query)
        .sort("created_at", 1)
        .skip(skip)
        .limit(limit)
    ):
        entry["_id"] = str(entry["_id"])
        entry["formatted_created_at"] = format_timestamp(entry.get("created_at"))
        registrations.append(entry)

    return jsonify(
        {
            "registrations": registrations,
            "total": total_count,
            "page": page,
            "limit": limit,
            "has_more": skip + len(registrations) < total_count,
        }
    )


@app.route("/admin", methods=["GET"])
@admin_required
def admin_dashboard():
    colleges, courses = get_form_options()
    db_connected = get_db() is not None
    
    # Get total count for display
    db = get_db()
    total_registrations = db.registrations.count_documents({}) if db is not None else 0

    return render_template(
        "admin_dashboard.html",
        hero=hero_story,
        colleges=colleges,
        courses=courses,
        db_connected=db_connected,
        total_registrations=total_registrations,
    )


@app.route("/admin/options", methods=["POST"])
@admin_required
def admin_update_options():
    option_type = request.form.get("option_type")
    action = request.form.get("action", "add")
    value = request.form.get("value", "").strip()

    colleges, courses = get_form_options()
    if option_type not in {"college", "course"} or not value:
        flash("Please provide a valid value.", "danger")
        return redirect(url_for("admin_dashboard"))

    target_list = list(colleges if option_type == "college" else courses)

    if action == "remove":
        if value in target_list:
            target_list.remove(value)
    else:
        if value not in target_list:
            target_list.append(value)

    if option_type == "college":
        colleges = target_list
    else:
        courses = target_list

    saved = save_form_options(colleges, courses)
    if saved:
        flash("Options updated successfully.", "success")
    else:
        db_unavailable_message()
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/options/reorder", methods=["POST"])
@admin_required
def admin_reorder_options():
    payload = request.get_json(silent=True) or {}
    option_type = payload.get("option_type")
    new_order = payload.get("order")

    if option_type not in {"college", "course"} or not isinstance(new_order, list):
        return jsonify({"success": False, "message": "Invalid payload"}), 400

    colleges, courses = get_form_options()

    def merge_order(original, desired):
        seen = []
        for value in desired:
            if value in original and value not in seen:
                seen.append(value)
        for value in original:
            if value not in seen:
                seen.append(value)
        return seen

    if option_type == "college":
        updated_colleges = merge_order(colleges, new_order)
        saved = save_form_options(updated_colleges, courses)
    else:
        updated_courses = merge_order(courses, new_order)
        saved = save_form_options(colleges, updated_courses)

    status_code = 200 if saved else 500
    return jsonify({"success": bool(saved)}), status_code


@app.route("/admin/registrations/<reg_id>/delete", methods=["POST"])
@admin_required
def admin_delete_registration(reg_id):
    db = get_db()
    if db is None:
        db_unavailable_message()
        return redirect(url_for("admin_dashboard"))

    try:
        object_id = ObjectId(reg_id)
    except InvalidId:
        flash("Invalid registration identifier.", "danger")
        return redirect(url_for("admin_dashboard"))

    try:
        result = db.registrations.delete_one({"_id": object_id})
    except errors.PyMongoError as exc:
        app.logger.error("Failed to delete registration: %s", exc)
        flash("Could not delete the registration. Please try again.", "danger")
        return redirect(url_for("admin_dashboard"))

    if result.deleted_count:
        flash("Registration removed permanently.", "success")
    else:
        flash("Registration was not found or already removed.", "warning")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/export/excel")
@admin_required
def export_excel():
    search_query = request.args.get("search", "").strip()
    college_filter = request.args.get("college", "").strip()
    registrations = fetch_registrations(
        search_query if search_query else None,
        college_filter if college_filter else None
    )
    if not registrations:
        flash("No registrations to export.", "warning")
        return redirect(url_for("admin_dashboard"))

    dataframe = pd.DataFrame(
        [
            {
                "Name": reg.get("name"),
                "College": reg.get("college"),
                "Course": reg.get("course"),
                "Role": reg.get("role") or reg.get("category", ""),
                "Phone": format_phone(reg.get("phone")),
                "Email": reg.get("email"),
                "Registered On": format_timestamp(reg.get("created_at")),
            }
            for reg in registrations
        ]
    )
    buffer = BytesIO()
    report_name = "Laksha Kantha Geetha Parayana Registration Sheet"
    
    # Determine start row based on whether there are filters
    has_filters = bool(college_filter or search_query)
    if has_filters:
        start_row = 3  # Title (1), Filter (2), Headers (3)
    else:
        start_row = 2  # Title (1), Headers (2)
    
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        dataframe.to_excel(
            writer, index=False, sheet_name="Registrations", startrow=start_row
        )
        worksheet = writer.sheets["Registrations"]

        # Title row
        worksheet.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=len(dataframe.columns),
        )
        title_cell = worksheet.cell(row=1, column=1)
        title_cell.value = report_name
        title_cell.font = Font(size=16, bold=True, color="1A237E")
        title_cell.alignment = Alignment(horizontal="center")

        # Filter row (only if filters exist)
        if has_filters:
            worksheet.merge_cells(
                start_row=2,
                start_column=1,
                end_row=2,
                end_column=len(dataframe.columns),
            )
            filter_cell = worksheet.cell(row=2, column=1)
            filter_parts = []
            if college_filter:
                filter_parts.append(f"College: {college_filter}")
            if search_query:
                filter_parts.append(f"Search: {search_query}")
            filter_cell.value = "\n".join(filter_parts)
            filter_cell.font = Font(size=12, italic=True, color="555555")
            filter_cell.alignment = Alignment(horizontal="center")
            header_row = 3
        else:
            header_row = 2
        header_fill = PatternFill("solid", fgColor="0B0A08")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )
        for col_idx, column in enumerate(dataframe.columns, start=1):
            cell = worksheet.cell(row=header_row, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        for idx, column in enumerate(dataframe.columns, start=1):
            max_len = len(column)
            column_series = dataframe[column].astype(str)
            if not column_series.empty:
                max_len = max(max_len, column_series.map(len).max())
            worksheet.column_dimensions[get_column_letter(idx)].width = max_len + 2

        for row in worksheet.iter_rows(
            min_row=header_row + 1,
            max_row=worksheet.max_row,
            min_col=1,
            max_col=len(dataframe.columns),
        ):
            for cell in row:
                cell.border = thin_border

    buffer.seek(0)
    filename = "registrations.xlsx"
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
    )


@app.route("/admin/export/pdf")
@admin_required
def export_pdf():
    search_query = request.args.get("search", "").strip()
    college_filter = request.args.get("college", "").strip()
    registrations = fetch_registrations(
        search_query if search_query else None,
        college_filter if college_filter else None
    )
    if not registrations:
        flash("No registrations to export.", "warning")
        return redirect(url_for("admin_dashboard"))

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    # Set consistent line width for borders (same on all pages)
    pdf.set_line_width(0.1)
    
    headers = ["Name", "College", "Course", "Role", "Phone", "Email", "Registered On"]
    col_widths = [32, 45, 25, 25, 28, 40, 35]
    usable_width = pdf.w - 2 * pdf.l_margin
    width_scale = usable_width / sum(col_widths)
    col_widths = [w * width_scale for w in col_widths]
    
    line_height = 6
    header_height = 8
    header_y = None

    def print_page_header():
        """Print page header (title, filter info) on every page."""
        pdf.set_font("Helvetica", "B", 16)
        pdf.cell(0, 10, "Laksha Kantha Geetha Parayana Registrations", ln=True)
        pdf.set_font("Helvetica", "", 11)
        # Filter info on separate line (if exists)
        if college_filter or search_query:
            filter_parts = []
            if college_filter:
                filter_parts.append(f"College: {college_filter}")
            if search_query:
                filter_parts.append(f"Search: {search_query}")
            pdf.cell(0, 8, "\n".join(filter_parts), ln=True)
        pdf.ln(6)
    
    def print_table_header():
        """Print table header row on current page."""
        nonlocal header_y
        pdf.set_font("Helvetica", "B", 11)
        pdf.set_fill_color(26, 35, 126)
        pdf.set_text_color(255, 255, 255)
        header_y = pdf.get_y()
        for header, width in zip(headers, col_widths):
            pdf.cell(width, header_height, header, border=1, align="C", fill=True)
        pdf.ln()
        # Reset font and text color for data rows (same as first page)
        pdf.set_font("Helvetica", size=10)
        pdf.set_text_color(40, 40, 40)

    def split_text(text, width):
        """Split text to fit within column width."""
        lines = pdf.multi_cell(width, line_height, str(text), split_only=True)
        return lines or [""]

    # First page
    pdf.add_page()
    print_page_header()
    print_table_header()

    pdf.set_font("Helvetica", size=10)

    for index, reg in enumerate(registrations):
        # Calculate row height first to check if it fits
        row = [
            reg.get("name", ""),
            reg.get("college", ""),
            reg.get("course", ""),
            reg.get("role") or reg.get("category", ""),
            format_phone(reg.get("phone", "")),
            reg.get("email", ""),
            format_timestamp(reg.get("created_at")),
        ]
        row_lines = [
            split_text(text, width) for text, width in zip(row, col_widths)
        ]
        max_lines = max(len(lines) for lines in row_lines)
        row_height = max_lines * line_height
        
        # Check if we need a new page (leave space for page header and table header)
        # Page header: 10px (title) + 8px (filter if exists) + 6px (spacing) = ~24px
        # Table header: 8px
        # Bottom margin: 15px
        # Total needed: ~24 + 8 + 15 = 47px minimum, plus row height
        space_needed = 47 + row_height
        if pdf.get_y() + space_needed > pdf.h - pdf.b_margin:
            pdf.add_page()
            # Reset line width for consistent borders (same as first page)
            pdf.set_line_width(0.1)
            print_page_header()
            print_table_header()
            # Ensure font is reset for data rows (same style as first page)
            pdf.set_font("Helvetica", size=10)
            pdf.set_text_color(40, 40, 40)

        if index % 2 == 0:
            pdf.set_fill_color(255, 255, 255)
        else:
            pdf.set_fill_color(245, 245, 245)

        y_start = pdf.get_y()
        x_pos = pdf.l_margin
        for width in col_widths:
            pdf.set_xy(x_pos, y_start)
            pdf.cell(width, row_height, "", border=1, fill=True)
            x_pos += width

        pdf.set_xy(pdf.l_margin, y_start)
        # Ensure font is set correctly for data (not bold, size 10) - same style on all pages
        pdf.set_font("Helvetica", size=10)
        pdf.set_text_color(40, 40, 40)
        for lines, width in zip(row_lines, col_widths):
            x = pdf.get_x()
            y = pdf.get_y()
            pdf.multi_cell(
                width,
                line_height,
                "\n".join(lines),
                border=0,
                align="L",
            )
            pdf.set_xy(x + width, y)
        pdf.set_xy(pdf.l_margin, y_start + row_height)

    pdf_data = pdf.output(dest="S")
    if isinstance(pdf_data, str):
        pdf_bytes = pdf_data.encode("latin1")
    else:
        pdf_bytes = bytes(pdf_data)
    buffer = BytesIO(pdf_bytes)
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name="registrations.pdf",
        mimetype="application/pdf",
    )


if __name__ == "__main__":
    app.run(debug=True,host='0.0.0.0', port=5002)